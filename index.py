"""
Relevance – Avaliação de Aderência
API serverless para deploy no Vercel (substitui o app Streamlit original).

Endpoints:
  POST /api/columns  -> lê o cabeçalho do xlsx e sugere as colunas (Ementa, Indexação, etc.)
  POST /api/process  -> roda a avaliação completa e devolve o xlsx de resultado (base64)

Observações importantes para o deploy:
  - Funções serverless da Vercel têm limite de tempo de execução (10s no plano Hobby,
    até 60s/300s em planos pagos, configurável em vercel.json -> functions.maxDuration).
    A etapa de consulta ao "Inteiro Teor" (download de PDFs da Câmara) pode ser lenta —
    ajuste maxDuration e/ou desative essa etapa se estourar o tempo no seu plano.
  - O corpo da requisição (upload do xlsx) tem limite de tamanho (~4.5MB por padrão
    na Vercel). Planilhas muito grandes podem precisar de outra estratégia de upload.
  - Não há mais st.session_state: cada requisição é isolada (serverless é stateless),
    então o cache de Inteiro Teor por URL dura apenas durante aquela requisição.
"""

import io
import re
import base64
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import numpy as np
import requests
from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from bs4 import BeautifulSoup
    BS4_OK = True
except ImportError:
    BS4_OK = False

app = Flask(__name__)

# ══════════════════════════════════════════════════════════════════════════
# NORMALIZAÇÃO E TERMOS (idêntico à lógica original)
# ══════════════════════════════════════════════════════════════════════════

def norm(t: str) -> str:
    t = (t or "").lower().strip()
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r"[^a-z0-9\s]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


_NOISE = {
    "de","da","do","das","dos","em","na","no","nas","nos","que","com","para",
    "por","uma","uns","um","as","os","a","o","e","ou","se","ao","aos","sobre",
    "entre","ate","apos","nao","sim","mas","mais","seu","sua","seus","suas",
    "este","esta","isso","aqui","ali","como","quando","onde","qual","quais",
    "ser","ter","foi","sao","via","lei","art","inc","par","num","nos","pelo",
    "pela","pelos","pelas","esse","essa","esses","essas","deste","desta","isto",
    "aquele","aquela","neste","nesta","numa","tambem","ainda","apenas","alem",
    "sendo","tendo","sido","seja","sejam","serao","sera","podem","pode",
    "todas","todos","todo","toda","cada","tipo","forma","outro","outra",
    "mesmo","mesma","proprio","propria","dado","dados","caso","casos",
    "rara","raras","raro","raros","nova","novo","novas","novos",
    "geral","gerais","social","sociais","civil","civis","vez","vezes",
}
_SIGLAS_OK = {"tea","bpc","pcd","sus","eca","loas","apae","cid","onu","oms"}


def _valid(v: str) -> bool:
    if not v or v in _NOISE:
        return False
    if len(v) <= 2:
        return False
    if len(v) == 3 and v not in _SIGLAS_OK:
        return False
    return True


def parse_terms(raw: str) -> list:
    """
    Extrai termos de busca:
      termo*         -> ('prefix', 'termo')
      "frase exata"  -> ('exact', 'frase exata')
      "prefix*"      -> ('prefix', 'prefix')
      palavra        -> ('word', 'palavra')
    """
    terms = []
    raw = raw or ""
    if not raw.strip():
        return terms

    QUOTE = r'["\u201c\u201d\u2018\u2019\u00ab\u00bb]'

    for m in re.finditer(QUOTE + r'([^"\u201c\u201d\u2018\u2019\u00ab\u00bb]+)' + QUOTE, raw):
        phrase = m.group(1).strip()
        if phrase.endswith('*'):
            prefix = norm(phrase[:-1]).strip()
            if len(prefix) >= 3:
                terms.append(('prefix', prefix))
        else:
            v = norm(phrase).strip()
            if not v:
                continue
            all_words = v.split()
            if not all_words:
                continue
            if len(all_words) == 1:
                if _valid(all_words[0]):
                    terms.append(('word', all_words[0]))
            else:
                terms.append(('exact', v))

    raw2 = re.sub(QUOTE + r'[^"\u201c\u201d\u2018\u2019\u00ab\u00bb]*' + QUOTE, ' ', raw)
    raw2 = re.sub(r'\b(or|and|not)\b', ' ', raw2, flags=re.IGNORECASE)
    raw2 = re.sub(r'[(),;]', ' ', raw2)

    tokens = raw2.split()
    merged = []
    i = 0
    while i < len(tokens):
        if i + 1 < len(tokens) and tokens[i + 1] == '*':
            merged.append(tokens[i] + '*')
            i += 2
        else:
            merged.append(tokens[i])
            i += 1

    for token in merged:
        if token.endswith('*'):
            prefix = norm(token[:-1]).strip()
            if len(prefix) >= 3:
                terms.append(('prefix', prefix))
        else:
            v = norm(token).strip()
            if _valid(v):
                terms.append(('word', v))

    seen = set()
    unique = []
    for t in terms:
        if t not in seen:
            seen.add(t)
            unique.append(t)
    return unique


def term_match(ttype: str, tval: str, tn: str, words: set) -> bool:
    if ttype == "exact":
        return bool(re.search(r"\b" + re.escape(tval) + r"\b", tn))
    elif ttype == "prefix":
        return any(w.startswith(tval) for w in words)
    else:
        if tval in words:
            return True
        if tval + "s" in words or tval + "es" in words:
            return True
        if tval.endswith("s") and len(tval) > 4 and tval[:-1] in words:
            return True
        if tval.endswith("es") and len(tval) > 5 and tval[:-2] in words:
            return True
        if len(tval) >= 6:
            return any(w.startswith(tval) for w in words)
        return False


def score_relevance(ementa: str, indexacao: str, theme: str, extra_raw: str) -> tuple:
    texto = norm(ementa + " " + indexacao)
    words = set(texto.split())

    theme_terms = parse_terms(theme)
    extra_terms = parse_terms(extra_raw)

    theme_vals = {v for _, v in theme_terms}
    extra_terms = [t for t in extra_terms if t[1] not in theme_vals]

    all_terms = theme_terms + extra_terms
    if not all_terms:
        return 5, "Nenhum tema/termo definido — proposição listada sem filtro de relevância."

    ht = sum(1 for tt, tv in theme_terms if term_match(tt, tv, texto, words))
    he = sum(1 for tt, tv in extra_terms if term_match(tt, tv, texto, words))

    if ht == 0 and he == 0:
        score = 1
    else:
        n_theme = len(theme_terms)
        n_extra = len(extra_terms)

        cov_t = ht / n_theme if n_theme > 0 else 0.0

        if n_extra == 0:
            if cov_t >= 0.70:
                score = 5
            elif cov_t >= 0.40:
                score = 4
            elif cov_t >= 0.15:
                score = 3
            else:
                score = 2
        else:
            if cov_t >= 0.70:
                base = 4
            elif cov_t >= 0.30:
                base = 3
            elif cov_t > 0.0:
                base = 2
            else:
                base = 1

            if he >= 5:
                boost = 2
            elif he >= 2:
                boost = 1
            elif he >= 1:
                boost = 1
            else:
                boost = 0

            if base == 1 and he >= 3:
                base = 2

            score = min(5, base + boost)

            if he >= 1 and score < 4:
                score = 4

    found_t = [tv for tt, tv in theme_terms if term_match(tt, tv, texto, words)][:5]
    found_e = [tv for tt, tv in extra_terms if term_match(tt, tv, texto, words)][:6]
    ft = ", ".join(found_t) or "nenhum"
    fe = ", ".join(found_e) or "nenhum"

    msgs = {
        5: f"Alta aderência. Tema: {ft}. Termos adicionais: {fe}.",
        4: f"Boa aderência. Tema: {ft}. Termos adicionais: {fe}.",
        3: f"Aderência moderada. Tema: {ft}. Termos adicionais: {fe}.",
        2: f"Baixa aderência. Tema: {ft}. Termos adicionais: {fe}.",
        1: "Sem aderência identificada. Nenhum termo do tema ou dos termos adicionais encontrado na ementa ou indexação.",
    }
    return score, msgs[score]


# ══════════════════════════════════════════════════════════════════════════
# INTEIRO TEOR — extração de hyperlinks e busca de texto
# ══════════════════════════════════════════════════════════════════════════

def _extract_pdf_text_search(content: bytes, search_terms: list) -> str:
    def _check(text, terms):
        tn = norm(text)
        return any(t in tn for t in terms)

    try:
        import pdfplumber
        parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for i, page in enumerate(pdf.pages):
                t = page.extract_text() or ""
                parts.append(t)
                combined = " ".join(parts)
                if search_terms and _check(combined, search_terms):
                    return combined
                if not search_terms and i >= 9:
                    break
        return " ".join(parts)
    except Exception:
        pass

    try:
        import pypdf
        parts = []
        reader = pypdf.PdfReader(io.BytesIO(content))
        for i, page in enumerate(reader.pages):
            t = page.extract_text() or ""
            parts.append(t)
            combined = " ".join(parts)
            if search_terms and _check(combined, search_terms):
                return combined
            if not search_terms and i >= 9:
                break
        return " ".join(parts)
    except Exception:
        return ""


def _fetch_teor(url: str, search_terms: list, cache: dict) -> str:
    if not url:
        return ""
    if url in cache:
        return cache[url]

    hdrs = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/124.0.0.0 Safari/537.36"),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "pt-BR,pt;q=0.9",
        "Referer": "https://www.camara.leg.br/",
    }
    text = ""
    try:
        resp = requests.get(url, headers=hdrs, timeout=20, stream=True)
        if resp.status_code == 200:
            ct = resp.headers.get("Content-Type", "").lower()
            content = b""
            is_pdf = "pdf" in ct

            for chunk in resp.iter_content(chunk_size=262144):
                content += chunk
                if not is_pdf and content[:4] == b"%PDF":
                    is_pdf = True
                if is_pdf and len(content) % 262144 < 32768:
                    partial_text = _extract_pdf_text_search(content, search_terms)
                    if partial_text and search_terms:
                        t_norm = norm(partial_text)
                        if any(t in t_norm for t in search_terms):
                            text = partial_text
                            break
                if len(content) >= 3_000_000:
                    break

            if not text:
                if is_pdf:
                    text = _extract_pdf_text_search(content, search_terms)
                elif BS4_OK:
                    soup = BeautifulSoup(content, "html.parser")
                    for tag in soup(["script", "style", "nav", "header", "footer", "noscript"]):
                        tag.decompose()
                    text = soup.get_text(separator=" ", strip=True)
                else:
                    text = content.decode("utf-8", errors="replace")
    except Exception:
        text = ""

    cache[url] = text
    return text


def extract_teor_links(file_bytes: bytes) -> dict:
    links = {}
    if not file_bytes:
        return links
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        teor_col = None
        for cell in ws[1]:
            if cell.value and "teor" in str(cell.value).lower():
                teor_col = cell.column
                break
        if teor_col is None:
            return links
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
            cell = row[teor_col - 1]
            url = None
            if cell.hyperlink:
                url = (getattr(cell.hyperlink, 'target', None)
                       or getattr(cell.hyperlink, 'display', None))
            elif cell.value and str(cell.value).startswith("http"):
                url = str(cell.value)
            if url:
                links[row_idx] = url
    except Exception:
        pass
    return links


def find_col(header, keywords):
    for i, h in enumerate(header):
        hn = norm(str(h))
        if any(kw in hn for kw in keywords):
            return i
    return None


def auto_detect_columns(header: list) -> dict:
    col_ementa = find_col(header, ["ementa"])
    col_indexa = find_col(header, ["indexa"])
    col_tema = find_col(header, ["tema"])
    col_teor = find_col(header, ["inteiro teor", "teor", "inteiroteor"])
    if col_ementa is None:
        col_ementa = find_col(header, ["descricao", "objeto", "assunto", "resumo"])
    if col_indexa is None:
        col_indexa = find_col(header, ["keywords", "palavras", "tag"])
    if col_ementa is None:
        col_ementa = 0
    return {
        "col_ementa": col_ementa,
        "col_indexa": col_indexa,
        "col_extra_text": col_tema,
        "col_teor": col_teor,
    }


# ══════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════

def build_output_excel(df_original: pd.DataFrame, scores: list, justifications: list) -> bytes:
    orig_headers = [str(v) if pd.notna(v) else "" for v in df_original.iloc[0]]
    new_headers = [orig_headers[0], "Índice de aderência (1-5)", "Justificativa"] + orig_headers[1:]
    wb = Workbook()
    ws = wb.active
    ws.title = "Avaliação"
    hfill = PatternFill("solid", fgColor="1A6FD4")
    hfont = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="C5D9F2")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    sfills = {
        1: PatternFill("solid", fgColor="FDDEDE"),
        2: PatternFill("solid", fgColor="FDE8CC"),
        3: PatternFill("solid", fgColor="FEF9C3"),
        4: PatternFill("solid", fgColor="D5F5E3"),
        5: PatternFill("solid", fgColor="D6EAF8"),
    }
    for ci, v in enumerate(new_headers, 1):
        c = ws.cell(row=1, column=ci, value=v)
        c.fill = hfill
        c.font = hfont
        c.alignment = center
        c.border = brd
    for rn, (score, justif) in enumerate(zip(scores, justifications), 2):
        orig = df_original.iloc[rn - 1]
        row_vals = (
            [orig.iloc[0] if pd.notna(orig.iloc[0]) else "", score, justif]
            + [(orig.iloc[j] if pd.notna(orig.iloc[j]) else "") for j in range(1, len(orig_headers))]
        )
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=rn, column=ci, value=v)
            c.border = brd
            c.alignment = wrap
            if ci == 2:
                c.fill = sfills.get(score, PatternFill())
                c.alignment = center
                c.font = Font(bold=True, name="Calibri")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    for i in range(4, len(new_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 30
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════

def _clean(v: str) -> str:
    v = "" if v is None else str(v)
    return "" if v.lower() in ("nan", "-", "none") else v


def _read_dataframe(file_bytes: bytes):
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
    if df_raw.shape[0] < 2:
        raise ValueError("A planilha precisa ter cabeçalho e ao menos uma linha de dados.")
    header = df_raw.iloc[0].tolist()
    data_rows = df_raw.iloc[1:].reset_index(drop=True)
    return df_raw, header, data_rows


def _optint(v):
    if v is None or v == "" or v == "null":
        return None
    return int(v)


# ══════════════════════════════════════════════════════════════════════════
# ROTAS
# ══════════════════════════════════════════════════════════════════════════

@app.route("/api/columns", methods=["POST"])
def api_columns():
    """Recebe o xlsx, devolve cabeçalho + colunas sugeridas + nº de links de Inteiro Teor."""
    f = request.files.get("file")
    if f is None:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400
    file_bytes = f.read()
    try:
        _, header, data_rows = _read_dataframe(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Erro ao ler o arquivo: {e}"}), 400

    teor_links = extract_teor_links(file_bytes)
    cols = auto_detect_columns(header)

    return jsonify({
        "header": [str(h) for h in header],
        "n_rows": int(len(data_rows)),
        "n_teor_links": len(teor_links),
        **cols,
    })


@app.route("/api/process", methods=["POST"])
def api_process():
    """Roda a avaliação completa e devolve os resultados + xlsx (base64)."""
    f = request.files.get("file")
    if f is None:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400
    file_bytes = f.read()

    main_theme = request.form.get("main_theme", "")
    extra_terms_raw = request.form.get("extra_terms", "")
    exclude_terms = request.form.get("exclude_terms", "")
    exclude_scope = request.form.get("exclude_scope", "Ementa e Indexação")
    fetch_teor = request.form.get("fetch_teor", "true").lower() == "true"

    col_ementa = _optint(request.form.get("col_ementa"))
    col_indexa = _optint(request.form.get("col_indexa"))
    col_extra_text = _optint(request.form.get("col_extra_text"))
    col_teor = _optint(request.form.get("col_teor"))

    try:
        df_raw, header, data_rows = _read_dataframe(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Erro ao ler o arquivo: {e}"}), 400

    if col_ementa is None:
        return jsonify({"error": "Selecione ao menos a coluna da Ementa."}), 400

    teor_links = extract_teor_links(file_bytes) if col_teor is not None else {}

    total = len(data_rows)
    scores_list, justif_list, rows_data = [], [], []

    for idx in range(total):
        row = data_rows.iloc[idx]
        ementa = _clean(row.iloc[col_ementa]) if col_ementa is not None else ""
        indexa = _clean(row.iloc[col_indexa]) if col_indexa is not None else ""
        extratx = _clean(row.iloc[col_extra_text]) if col_extra_text is not None else ""
        indexa_full = " ".join(filter(None, [indexa, extratx]))
        rows_data.append((ementa, indexa_full))
        s, j = score_relevance(ementa, indexa_full, main_theme, extra_terms_raw)
        scores_list.append(s)
        justif_list.append(j)

    # Passo 2: consultar Inteiro Teor apenas das que ficaram com score = 1
    teor_checked = 0
    teor_available = None
    if fetch_teor and col_teor is not None and teor_links:
        score1_idxs = [i for i, s in enumerate(scores_list) if s == 1 and i in teor_links]
        if score1_idxs:
            tt_all = parse_terms(main_theme) + parse_terms(extra_terms_raw)
            search_terms = [norm(tv) for _, tv in tt_all if len(tv) >= 3]

            first_url = teor_links[score1_idxs[0]]
            try:
                hdrs_test = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
                r_test = requests.get(first_url, headers=hdrs_test, timeout=8, stream=True)
                r_test.close()
                teor_available = r_test.status_code == 200
            except Exception:
                teor_available = False

            if teor_available:
                cache = {}
                links_score1 = {i: teor_links[i] for i in score1_idxs}
                with ThreadPoolExecutor(max_workers=8) as executor:
                    futures = {
                        executor.submit(_fetch_teor, url, search_terms, cache): idx_f
                        for idx_f, url in links_score1.items()
                    }
                    for future in as_completed(futures):
                        idx_f = futures[future]
                        teor = future.result()
                        teor_checked += 1
                        if teor:
                            s_teor, _ = score_relevance(teor, "", main_theme, extra_terms_raw)
                            if s_teor > 1:
                                scores_list[idx_f] = 2
                                justif_list[idx_f] = "Aderência encontrada no Inteiro Teor (não consta na ementa/indexação)."

    # Passo 3: termos de exclusão
    for idx in range(total):
        ementa, indexa_full = rows_data[idx]
        s = scores_list[idx]
        j = justif_list[idx]
        if exclude_terms.strip() and s > 1:
            excl_terms = parse_terms(exclude_terms)
            if exclude_scope == "Somente Ementa":
                texto_excl = norm(ementa)
            elif exclude_scope == "Somente Indexação":
                texto_excl = norm(indexa_full)
            else:
                texto_excl = norm(ementa + " " + indexa_full)
            words_excl = set(texto_excl.split())
            excl_hits = [tv for tp, tv in excl_terms if term_match(tp, tv, texto_excl, words_excl)]
            if excl_hits:
                s = 1
                j = f"Excluído ({exclude_scope.lower()}) — termo encontrado: {', '.join(excl_hits[:5])}."
        scores_list[idx] = s
        justif_list[idx] = j

    excel_bytes = build_output_excel(df_raw, scores_list, justif_list)
    excel_b64 = base64.b64encode(excel_bytes).decode("ascii")

    counts = {i: scores_list.count(i) for i in range(1, 6)}
    avg = float(np.mean(scores_list)) if scores_list else 0.0

    return jsonify({
        "propositions": [str(v) if pd.notna(v) else "" for v in data_rows.iloc[:, 0].tolist()],
        "scores": scores_list,
        "justifications": justif_list,
        "counts": counts,
        "average": avg,
        "n_teor_links": len(teor_links),
        "teor_checked": teor_checked,
        "teor_available": teor_available,
        "excel_base64": excel_b64,
    })


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


# Entry point local (não usado na Vercel, útil para testar com `python api/index.py`)
if __name__ == "__main__":
    app.run(debug=True, port=5000)
